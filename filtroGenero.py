import time
start=time.time()
import xlsxwriter
from openpyxl import load_workbook
#Comentario para crear cambios en el documento

print("\n|||----------------------  PROGRAMA DE FILTRADO DE REGISTROS CLÍNICOS  ------------------------|°|\n")
print("El programa inició a las: (hh:mm:ss):",time.strftime("%H:%M:%S"),"\n")
print("Abriendo el documento de Excel, espere por favor.")

workbook1 = xlsxwriter.Workbook("./regFiltrados.xlsx", {'constant_memory': True})#El libro nuevo
hoja=workbook1.add_worksheet("Sheet1")#La hoja de la base de datos 

# worksheet1 = load_workbook("./base_datos/HCL 2017-2018.xlsx")
# hoja1=worksheet1["DGExport"]

worksheet1 = load_workbook("./debuggPrograma.xlsx")#La hoja de la que se extraen los datos
hoja1=worksheet1["Sheet1"]#La hoja del nuevo documento .xlsx

#Muestra información de las filas y columnas de la hoja cargada
maxFila=hoja1.max_row
maxColumna=hoja1.max_column

#Lista que contendrá todos los datos de la fila de identificación por documento
k=[]
#lista de las filas donde se encontraron vacíos
l=[]

# rango=hoja1["I2":"I76700"]
rango=hoja1["I2":"I29"]

#------------------   LISTA DE DATOS       ----------------#

for i in rango:
    for j in i:
        nombre=j.value
        k.append(nombre)   #CREA EL VECTOR k QUE SERVIRÁ PARA REALIZAR LAS COMPARACIONES

print("La cantidad de registros en la base de datos es:",len(k),"\n")#Imprime la cantidad de datos en la lista k
#print(k)
#-----------AUXILIARES DEL CICLO DE COMPARACIONES--------#

contador=0#Es el índice que recorre el vector
nuevoCont=1#cuenta las veces que se ingresa determinado número de documento en la nueva hoja
filaNuevo=0#cuenta las filas para la nueva hoja de registros
interruptor=0#Permite agregar el primer dato de la hoja que es omitido por la forma en que está diseñado el programa
cantMujeres=0
cantHombres=0
colNueva=0
totalPersonas=0
nuevoIndice=0
numCitas=1
imprimio=0

#----CICLO DE COMPARACIONES HASTA QUE SE HAYAN ELIMINADO TODOS LOS nombreES REPETIDOS----#
                                #contador<=len(k)+1; len(k)>0
while numCitas==1:
    numCitas=int(input("Ingrese la cantidad mínima de registros, debe ser mayor a uno (1): "))
    if numCitas==1:
        print("\nDebe ingresar una cantidad mayor a uno (1).\n")
# if numCitas ==1:
#     pass
# else:
#     nuevoCont+=1
print("\nEl programa está separando en un nuevo documento de excel los datos sin repetir de las mujeres en la base de datos")
print("\nprocesando los datos...")
print("\n puede tomar algunos minutos")

lectorTemp=hoja1.cell(row=contador+1,column=1)#La menor de las filas es 1
registroTemp=lectorTemp.value
hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el nombre de la variable "fecha del registro de atención"
colNueva+=1

lectorTemp=hoja1.cell(row=contador+1,column=9)
registroTemp=lectorTemp.value
hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el nombre de la variable "doc_ident" a la hoja nueva(I1)
colNueva+=1

lectorTemp=hoja1.cell(row=contador+1,column=15)
registroTemp=lectorTemp.value
hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el nombre de la variable "DIRECCIÓN" a la hoja nueva(I1)
colNueva+=1

lectorTemp=hoja1.cell(row=contador+1,column=27)
registroTemp=lectorTemp.value
hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el nombre de la variable "Enfermedad_Actual" a la hoja nueva(I1)
colNueva+=1

lectorTemp=hoja1.cell(row=contador+1,column=40)
registroTemp=lectorTemp.value
hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el nombre de la variable "Dolor_Anginoso_Ejercicio" a la hoja nueva(I1)
colNueva+=1

lectorTemp=hoja1.cell(row=contador+1,column=43)
registroTemp=lectorTemp.value
hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el nombre de la variable "RS_Ortopnea" a la hoja nueva(I1)
filaNuevo+=1
colNueva=0

while (contador<=(len(k)-2)):
    if (k[contador]!=k[contador+1]):    #SI LOS NÚMEROS SON DIFERENTES EN LA LISTA K
        nuevoCont=1
        interruptor=0
        nuevoIndice=0
        d3 = hoja1.cell(row = contador+2, column = 20)
        v3=d3.value
        if v3=="MASCULINO":
            cantHombres+=1 
        else:#v3=="FEMENINO"
            cantMujeres+=1
        contador+=1
    else:                    #SI LOS NÚMEROS SON IGUALES EN LA LISTA K
        nuevoCont+=1
        if  nuevoCont>=numCitas:
            copiaCont=contador
            if interruptor<1:
                if numCitas>2:
                    copiaCont-=(nuevoCont-2)
                else:
                    pass
                while nuevoIndice<nuevoCont:
                    lectorTemp=hoja1.cell(row=copiaCont+2,column=15)
                    registroTemp=lectorTemp.value
                    celdaDir=registroTemp

                    lectorTemp=hoja1.cell(row=copiaCont+2,column=27)#Agrega la variable "EnfermedadActual"(AA)
                    registroTemp=lectorTemp.value
                    celdaEnferm=registroTemp
                    
                    lectorTemp1=hoja1.cell(row=copiaCont+2,column=40)#Agrega la variable "dolorAnginosoEjercicio"(AN)
                    registroTemp=lectorTemp1.value
                    dolAngValid=registroTemp
                    
                    lectorTemp2=hoja1.cell(row=copiaCont+2,column=43)#Toma el valor de la columna "RS_ORTOPNEA"
                    registroTemp=lectorTemp2.value
                    ortopnValid=registroTemp


                    #VERIFICA QUE NO HAYAN CELDAS VACÍAS EN EL NUEVO DOCUMENTO
                    if celdaEnferm==None or dolAngValid==None or ortopnValid==None or celdaDir==None:
                        # hoja.set_row(filaNuevo, None, None, {'hidden': True})
                        # print("celda vacía en fila: ",filaNuevo+1)
                        #vacios+=1
                        filaNuevo-=1
                        copiaFila=copiaCont+2
                        l.append(copiaFila)
                        #copiaFila+=1
                        copiaCont+=1
                        #nuevoCont-=1
                    else:
                        imprimio+=1
                        lectorTemp=hoja1.cell(row=copiaCont+2,column=1)
                        registroTemp=lectorTemp.value
                        hoja.write(filaNuevo,colNueva,registroTemp)#Agrega la fecha del registro de atención(A)
                        colNueva+=1
                        
                        lectorTemp=hoja1.cell(row=copiaCont+2,column=9)
                        registroTemp=lectorTemp.value
                        hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el doc_ident a la hoja nueva(I)
                        colNueva+=1

                        lectorTemp=hoja1.cell(row=copiaCont+2,column=15)
                        registroTemp=lectorTemp.value
                        hoja.write(filaNuevo,colNueva,registroTemp)#Agrega la dirección a la hoja nueva(O)
                        colNueva+=1
                        
                        lectorTemp=hoja1.cell(row=copiaCont+2,column=27)#Agrega la variable "EnfermedadActual"(AA)
                        registroTemp=lectorTemp.value
                        hoja.write(filaNuevo,colNueva,registroTemp )
                        colNueva+=1
                        
                        lectorTemp=hoja1.cell(row=copiaCont+2,column=40)#Agrega la variable "dolorAnginosoEjercicio"(AN)
                        registroTemp=lectorTemp.value
                        hoja.write(filaNuevo,colNueva,registroTemp )
                        colNueva+=1
                        
                        lectorTemp=hoja1.cell(row=copiaCont+2,column=43)#Agrega la variable "RS_ORTOPNEA"(AQ)
                        registroTemp=lectorTemp.value
                        hoja.write(filaNuevo,colNueva,registroTemp )
                        colNueva=0
                        copiaCont+=1
                    filaNuevo+=1 
                    nuevoIndice+=1
                    
                #contador=copiaCont+nuevoCont-1
                interruptor=1
                contador+=1
            else:#interruptor>1

                lectorTemp3=hoja1.cell(row=copiaCont+3,column=15)
                registroTemp=lectorTemp3.value
                celdaDir=registroTemp

                lectorTemp3=hoja1.cell(row=copiaCont+3,column=27)
                registroTemp=lectorTemp3.value
                celdaEnferm1=registroTemp
                
                lectorTemp4=hoja1.cell(row=copiaCont+3,column=40)
                registroTemp=lectorTemp4.value
                dolAngValid1=registroTemp
                
                lectorTemp5=hoja1.cell(row=copiaCont+3,column=43)
                registroTemp=lectorTemp5.value
                ortopnValid1=registroTemp
                if celdaEnferm1==None or dolAngValid1==None or ortopnValid1==None or celdaDir==None:
                    #     hoja.set_row(filaNuevo, None, None, {'hidden': True})
                    filaNuevo-=1
                    copiaFila=copiaCont+3
                    l.append(copiaFila)
                    # nuevoCont-=1
                    #copiaFila+=1
                    #pass
                else:
                    imprimio+=1
                    lectorTemp=hoja1.cell(row=copiaCont+3,column=1)
                    registroTemp=lectorTemp.value
                    hoja.write(filaNuevo,colNueva,registroTemp)#Agrega la fecha del registro de atención(A)
                    colNueva+=1
                    
                    lectorTemp=hoja1.cell(row=copiaCont+3,column=9)
                    registroTemp=lectorTemp.value
                    hoja.write(filaNuevo,colNueva,registroTemp)#Agrega el doc_ident a la hoja nueva(I)
                    colNueva+=1

                    lectorTemp=hoja1.cell(row=copiaCont+3,column=15)
                    registroTemp=lectorTemp.value
                    hoja.write(filaNuevo,colNueva,registroTemp)#Agrega la dirección a la hoja nueva(I)
                    colNueva+=1
                    
                    lectorTemp=hoja1.cell(row=copiaCont+3,column=27)#Agrega la variable "EnfermedadActual"(AA)
                    registroTemp=lectorTemp.value
                    hoja.write(filaNuevo,colNueva,registroTemp )
                    colNueva+=1
                    
                    lectorTemp=hoja1.cell(row=copiaCont+3,column=40)#Agrega la variable "dolorAnginosoEjercicio"(AN)
                    registroTemp=lectorTemp.value
                    hoja.write(filaNuevo,colNueva,registroTemp )
                    colNueva+=1
                    
                    lectorTemp=hoja1.cell(row=copiaCont+3,column=43)#Agrega la variable "RS_ORTOPNEA"(AQ)
                    registroTemp=lectorTemp.value
                    hoja.write(filaNuevo,colNueva,registroTemp )
                    colNueva=0
                filaNuevo+=1    
                contador+=1
        else:#Si nuevoCont<numCitas
            contador+=1
    #Si el índice de la lista es igual al penúltimo elemento en la lista K
    if (contador==(len(k)-2)):
        d3 = hoja1.cell(row = contador+3, column = 20)
        v3=d3.value
        if v3=="MASCULINO":
            cantHombres+=1
        else:
            cantMujeres+=1

print("\nlas filas omitidas fueron las número:",l[:])
workbook1.close()

print("La cantidad de mujeres es:",cantMujeres)
print("La cantidad de hombres es:",cantHombres)
totalPersonas=cantHombres+cantMujeres
print("Hay",totalPersonas,"personas.")

porcentajeHombres=round((cantHombres/totalPersonas)*100,3)
porcentajeMujeres=round(((cantMujeres)/totalPersonas)*100,3)
print("se omitieron:",len(l),"filas")

print("El porcentaje de Hombres en la base de datos es:",porcentajeHombres)
print("El porcentaje de Mujeres en la base de datos es:",porcentajeMujeres)

end=time.time()
tiempo=format(end-start)
floatTiempo=float(tiempo)
minutos=((floatTiempo)/60)
minf=int(minutos//1)
segundos=((minutos%1)*60)
segf=int(segundos)
miles=int((segundos%1)*1000)

print("\nEl tiempo de ejecución del programa fue de:", minf,"minutos", segf,"segundos", miles,"milisegundos")